import json
import os
from openpyxl import load_workbook, Workbook

def save_txt(path, data):
     with open(path, 'w', encoding='utf8') as f:
         f.write(data)

def read_txt(path):
    with open(path, 'r', encoding='utf8') as f:
            cont = f.read()
    return cont

def read_json(path):
    with open(path, 'r', encoding='utf8') as f:
        cont = json.loads(f.read())
    return cont


def save_json(path, data):
    with open(path, 'w', encoding='utf8') as f:
         f.write(json.dumps(data, ensure_ascii=False))

def is_exist_path(file_path):
    if not os.path.exists(file_path):
        flag = input(f'{file_path}不存在，是否创建')
        if flag: # 仅什么都不输入才会创立
            return False
        os.makedirs(file_path)
        print('创建成功')
    return True

def is_file_in_folder(folder_path, file_name):
    return file_name in os.listdir(folder_path)

def save_excel(path, data, firstLine=[]):
    wb = Workbook()
    ws = wb.create_sheet()
    if firstLine:
        ws.append(firstLine)

    for cont in data:
        ws.append(cont)

    wb.save(path)


def read_excel(path, sheet=None):
    wb = load_workbook(path)
    if sheet:
        ws = wb[sheet]
    else:
        ws = wb.active
    cont = ws.rows
    return [[j.value for j in i] for i in cont]